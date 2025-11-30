from openpyxl import load_workbook

# Tải (load) file Excel
wb = load_workbook('demo.xlsx')

# In ra tên tất cả các sheet
print("Tên các sheets có trong workbook:")
print(wb.sheetnames)

# Chọn sheet đầu tiên (chỉ mục 0)
ws = wb[wb.sheetnames[0]]

# Duyệt qua các giá trị trong sheet (row by row)
print("\nGiá trị đọc được từ sheet đầu tiên:")
for row in ws.values:
    # Duyệt qua từng ô (value) trong hàng
    for value in row:
        # In giá trị của ô, cách nhau bằng tab
        print(value, "\t", end='')
    # Xuống dòng sau khi in hết các giá trị của một hàng
    print("")