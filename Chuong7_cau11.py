from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

FILE_NAME = 'nhanvien.xlsx'

# --- 1. Cấu hình và Khởi tạo File Excel ---

def khoi_tao_file():
    """Tạo file Excel nếu chưa tồn tại, thiết lập tiêu đề cột."""
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhSachNhanVien"
        
        # Thiết lập tiêu đề cột
        headers = ['STT', 'Mã', 'Tên', 'Tuổi']
        ws.append(headers)
        
        # Thiết lập độ rộng cột cho dễ nhìn
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        
        wb.save(FILE_NAME)
        print(f"Đã tạo file {FILE_NAME} và thiết lập tiêu đề.")

# --- 2. Chức năng Lưu/Thêm Nhân Viên vào Excel ---

def luu_nhan_vien():
    """Nhập thông tin nhân viên mới và ghi vào hàng cuối cùng của Excel."""
    khoi_tao_file()
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    
    # Lấy STT mới
    stt_moi = ws.max_row
    
    ma_nv = input("Nhập Mã NV (ví dụ: NV7): ").strip()
    ten_nv = input("Nhập Tên NV: ").strip()
    
    while True:
        try:
            tuoi_nv = int(input("Nhập Tuổi: "))
            if tuoi_nv <= 0:
                print("Tuổi phải lớn hơn 0.")
                continue
            break
        except ValueError:
            print("Tuổi không hợp lệ. Vui lòng nhập số nguyên.")
            
    # Thêm dữ liệu vào hàng tiếp theo
    new_row = [stt_moi, ma_nv, ten_nv, tuoi_nv]
    ws.append(new_row)
    
    wb.save(FILE_NAME)
    print(f"Đã thêm nhân viên {ten_nv} ({ma_nv}) vào file Excel.")

# --- 3. Chức năng Đọc và Hiển thị Danh sách Nhân Viên ---

def doc_danh_sach(ws):
    """Đọc dữ liệu từ sheet và trả về List các Dictionary."""
    
    # Bỏ qua hàng tiêu đề
    data = []
    # Duyệt qua các hàng, bắt đầu từ hàng thứ 2
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[1]: # Kiểm tra nếu cột Mã (chỉ mục 1) có dữ liệu
            data.append({
                'STT': row[0],
                'Ma': row[1],
                'Ten': row[2],
                'Tuoi': row[3]
            })
    return data

def hien_thi_nhan_vien(nv_list):
    """In danh sách nhân viên ra console."""
    if not nv_list:
        print("Danh sách nhân viên rỗng.")
        return
        
    print("\n-------------------------------------------")
    print("STT | Mã\t| Tên\t\t| Tuổi")
    print("-------------------------------------------")
    for nv in nv_list:
        print(f"{nv['STT'] if nv['STT'] is not None else '':<3} | {nv['Ma']}\t| {nv['Ten']:<10}\t| {nv['Tuoi']}")
    print("-------------------------------------------")

# --- 4. Chức năng Sắp xếp Nhân Viên theo Tuổi ---

def sap_xep_theo_tuoi():
    """Đọc, sắp xếp theo Tuổi (tăng dần), và ghi đè lại file Excel."""
    khoi_tao_file()
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    
    # Đọc dữ liệu hiện tại
    nv_list = doc_danh_sach(ws)
    if not nv_list:
        print("Không có dữ liệu để sắp xếp.")
        return
        
    # Sắp xếp theo cột 'Tuoi' (tăng dần)
    # Hàm sort() sẽ sắp xếp tại chỗ.
    nv_list.sort(key=lambda x: int(x['Tuoi']), reverse=False)

    # Ghi lại toàn bộ dữ liệu (bắt đầu từ hàng 2)
    # Xóa dữ liệu cũ (từ hàng 2 đến hết)
    ws.delete_rows(2, ws.max_row) 
    
    # Ghi dữ liệu đã sắp xếp
    for i, nv in enumerate(nv_list):
        # Đặt STT theo thứ tự mới sau khi sắp xếp
        new_row = [i + 1, nv['Ma'], nv['Ten'], nv['Tuoi']]
        ws.append(new_row)
        
    wb.save(FILE_NAME)
    print("\nĐã sắp xếp nhân viên theo Tuổi (tăng dần) và lưu lại vào file.")
    # Hiển thị kết quả sau khi sắp xếp
    hien_thi_nhan_vien(doc_danh_sach(ws))


# --- CHƯƠNG TRÌNH CHÍNH ---

import os
def ChayQLNhanVien():
    print("\n\n============ QUẢN LÝ NHÂN VIÊN (EXCEL) ============")
    while True:
        print("\n1. Xem DS | 2. Thêm NV | 3. Sắp xếp (Tuổi) | 0. Thoát")
        chon = input("Chọn chức năng: ")
        
        if chon == '1':
            khoi_tao_file()
            wb = load_workbook(FILE_NAME)
            hien_thi_nhan_vien(doc_danh_sach(wb.active))
        elif chon == '2':
            luu_nhan_vien()
        elif chon == '3':
            sap_xep_theo_tuoi()
        elif chon == '0':
            print("Kết thúc chương trình Quản Lý Nhân Viên.")
            break
        else:
            print("Chọn không hợp lệ.")

# Chạy chương trình
ChayQLNhanVien()